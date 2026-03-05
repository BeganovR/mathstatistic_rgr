"""
Генерация Excel-файла со статистической обработкой выборки X (N=121..240).
Запуск: python generate_excel.py
"""

import math
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, numbers
)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Исходные данные
# ---------------------------------------------------------------------------

N_START = 121
X_VALUES = [
    172, 173, 173, 165, 167, 173, 184, 163, 179, 161, 162, 158,
    171, 177, 164, 166, 171, 174, 170, 174, 169, 174, 169, 175,
    167, 172, 168, 163, 168, 161, 173, 164, 167, 164, 173, 176,
    172, 167, 173, 161, 171, 169, 161, 170, 174, 168, 164, 170,
    164, 162, 166, 172, 169, 169, 163, 178, 166, 168, 168, 180,
    163, 165, 163, 158, 171, 175, 170, 165, 184, 169, 167, 167,
    179, 165, 173, 161, 166, 164, 159, 175, 169, 172, 172, 167,
    160, 156, 161, 174, 167, 174, 167, 168, 168, 167, 167, 171,
    168, 162, 174, 173, 173, 165, 167, 172, 176, 174, 171, 169,
    161, 173, 170, 176, 171, 166, 171, 167, 156, 167, 166, 167,
]

n = len(X_VALUES)  # 120

# ---------------------------------------------------------------------------
# Стилевые утилиты
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
SUBHEADER_FILL = PatternFill("solid", fgColor="BDD7EE")
SUM_FILL = PatternFill("solid", fgColor="DDEBF7")
LIGHT_FILL = PatternFill("solid", fgColor="EBF3FB")

HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
SUBHEADER_FONT = Font(bold=True, name="Calibri", size=11)
NORMAL_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(bold=True, name="Calibri", size=11)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")

THIN = Side(border_style="thin", color="000000")
MEDIUM = Side(border_style="medium", color="000000")

THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MEDIUM_BORDER = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)


def header_border():
    return Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)


def cell_border():
    return Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(cell, text=None, fill=None):
    if text is not None:
        cell.value = text
    cell.font = HEADER_FONT
    cell.fill = fill or HEADER_FILL
    cell.alignment = CENTER
    cell.border = cell_border()


def style_subheader(cell, text=None):
    if text is not None:
        cell.value = text
    cell.font = SUBHEADER_FONT
    cell.fill = SUBHEADER_FILL
    cell.alignment = CENTER
    cell.border = cell_border()


def style_data(cell, value=None, align=CENTER):
    if value is not None:
        cell.value = value
    cell.font = NORMAL_FONT
    cell.alignment = align
    cell.border = cell_border()


def style_sum(cell, value=None):
    if value is not None:
        cell.value = value
    cell.font = BOLD_FONT
    cell.fill = SUM_FILL
    cell.alignment = CENTER
    cell.border = cell_border()


def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


# ---------------------------------------------------------------------------
# Статистические вычисления
# ---------------------------------------------------------------------------

x_sorted = sorted(X_VALUES)
x_min = x_sorted[0]
x_max = x_sorted[-1]
R = x_max - x_min

# Шаг интервала
h_raw = R / (1 + 3.32 * math.log10(n))
h = math.ceil(h_raw)  # Округляем вверх до целого

# Число интервалов
k_intervals = math.ceil(R / h)

# Статистический ряд (точечный)
from collections import Counter

freq_counter = Counter(X_VALUES)
unique_vals = sorted(freq_counter.keys())
freqs = [freq_counter[v] for v in unique_vals]
rel_freqs = [f / n for f in freqs]

# Интервальный ряд
intervals = []
a = x_min
for i in range(k_intervals):
    b = a + h
    # Для последнего интервала включаем правую границу
    if i == k_intervals - 1:
        cnt = sum(1 for x in X_VALUES if a <= x <= b)
    else:
        cnt = sum(1 for x in X_VALUES if a <= x < b)
    mid = (a + b) / 2
    w = cnt / n
    intervals.append({
        "a": a, "b": b,
        "mid": mid,
        "k": cnt,
        "w": w,
        "density": w / h,
    })
    a = b

# Мода (Mo) по интервальному ряду
modal_idx = max(range(len(intervals)), key=lambda i: intervals[i]["k"])
modal = intervals[modal_idx]
if modal_idx > 0:
    k_prev = intervals[modal_idx - 1]["k"]
else:
    k_prev = 0
if modal_idx < len(intervals) - 1:
    k_next = intervals[modal_idx + 1]["k"]
else:
    k_next = 0
k_m = modal["k"]
denom_mo = (k_m - k_prev) + (k_m - k_next)
Mo = modal["a"] + h * (k_m - k_prev) / denom_mo if denom_mo != 0 else modal["mid"]

# Медиана (Me) по интервальному ряду
half_n = n / 2
cumsum = 0
Me = None
for iv in intervals:
    if cumsum + iv["k"] >= half_n:
        Me = iv["a"] + h * (half_n - cumsum) / iv["k"]
        break
    cumsum += iv["k"]

# Начальные моменты
alpha1 = sum(v * w for v, w in zip(unique_vals, rel_freqs))
alpha2 = sum(v**2 * w for v, w in zip(unique_vals, rel_freqs))
alpha3 = sum(v**3 * w for v, w in zip(unique_vals, rel_freqs))
alpha4 = sum(v**4 * w for v, w in zip(unique_vals, rel_freqs))

# Характеристики выборки
x_mean = alpha1
D_v = alpha2 - alpha1**2
sigma_v = math.sqrt(D_v)

beta2 = D_v
beta3 = alpha3 - 3 * alpha1 * alpha2 + 2 * alpha1**3
beta4 = alpha4 - 4 * alpha1 * alpha3 + 6 * alpha1**2 * alpha2 - 3 * alpha1**4

A = beta3 / sigma_v**3
E = beta4 / sigma_v**4 - 3

# Точечные оценки (метод статистических оценок)
S2 = (n / (n - 1)) * D_v  # исправленная дисперсия
S = math.sqrt(S2)

# Интервальные оценки (γ=0.95)
gamma = 0.95
# t = квантиль нормального распределения для двустороннего интервала при γ=0.95 (z₀.₉₇₅ = 1.96)
t = 1.96
# q = коэффициент для дов. интервала σ при n=120 и γ=0.95 (из таблицы: q ≈ 0.143)
q = 0.143

delta = t * S / math.sqrt(n)
mean_low = x_mean - delta
mean_high = x_mean + delta

sigma_low = S * (1 - q)
sigma_high = S * (1 + q)

# Эмпирическая функция распределения
cdf_vals = []
cumsum = 0
for v in unique_vals:
    cumsum += freq_counter[v]
    cdf_vals.append(cumsum / n)

# ---------------------------------------------------------------------------
# Создание книги
# ---------------------------------------------------------------------------

wb = Workbook()

# ---------------------------------------------------------------------------
# Лист 1: Исходные данные
# ---------------------------------------------------------------------------

ws1 = wb.active
ws1.title = "Исходные данные"

ws1.merge_cells("A1:B1")
c = ws1["A1"]
c.value = "Исходные данные выборки X"
c.font = Font(bold=True, size=14, name="Calibri")
c.alignment = CENTER

style_header(ws1["A2"], "N")
style_header(ws1["B2"], "X")
set_col_width(ws1, 1, 8)
set_col_width(ws1, 2, 10)
ws1.row_dimensions[2].height = 20

for i, x in enumerate(X_VALUES):
    row = i + 3
    style_data(ws1.cell(row, 1), N_START + i)
    style_data(ws1.cell(row, 2), x)
    if i % 2 == 0:
        ws1.cell(row, 1).fill = LIGHT_FILL
        ws1.cell(row, 2).fill = LIGHT_FILL

# ---------------------------------------------------------------------------
# Лист 2: Вариационный ряд
# ---------------------------------------------------------------------------

ws2 = wb.create_sheet("Вариационный ряд")

ws2.merge_cells("A1:B1")
c = ws2["A1"]
c.value = "Вариационный ряд (X, отсортированный по возрастанию)"
c.font = Font(bold=True, size=14, name="Calibri")
c.alignment = CENTER

style_header(ws2["A2"], "№")
style_header(ws2["B2"], "X (вар. ряд)")
set_col_width(ws2, 1, 8)
set_col_width(ws2, 2, 14)

for i, x in enumerate(x_sorted):
    row = i + 3
    style_data(ws2.cell(row, 1), i + 1)
    style_data(ws2.cell(row, 2), x)
    if i % 2 == 0:
        ws2.cell(row, 1).fill = LIGHT_FILL
        ws2.cell(row, 2).fill = LIGHT_FILL

# ---------------------------------------------------------------------------
# Лист 3: Статистический ряд (Таблица 1)
# ---------------------------------------------------------------------------

ws3 = wb.create_sheet("Стат. ряд")

ws3.merge_cells("A1:D1")
c = ws3["A1"]
c.value = "Таблица 1. Статистический (точечный) ряд"
c.font = Font(bold=True, size=14, name="Calibri")
c.alignment = CENTER

headers3 = ["xᵢ", "kᵢ (частота)", "wᵢ = kᵢ/n", "Накопл. частота"]
for col, h_text in enumerate(headers3, 1):
    style_header(ws3.cell(2, col), h_text)
    set_col_width(ws3, col, 16)

cum_k = 0
for i, (v, f, w) in enumerate(zip(unique_vals, freqs, rel_freqs)):
    row = i + 3
    cum_k += f
    style_data(ws3.cell(row, 1), v)
    style_data(ws3.cell(row, 2), f)
    style_data(ws3.cell(row, 3), round(w, 6))
    style_data(ws3.cell(row, 4), cum_k)
    if i % 2 == 0:
        for col in range(1, 5):
            ws3.cell(row, col).fill = LIGHT_FILL

sum_row = len(unique_vals) + 3
style_sum(ws3.cell(sum_row, 1), "Σ")
style_sum(ws3.cell(sum_row, 2), n)
style_sum(ws3.cell(sum_row, 3), round(sum(rel_freqs), 6))
style_sum(ws3.cell(sum_row, 4), "")

# Дополнительные параметры
info_row = sum_row + 2
ws3.cell(info_row, 1).value = "n ="
ws3.cell(info_row, 1).font = BOLD_FONT
ws3.cell(info_row, 2).value = n
ws3.cell(info_row, 2).font = NORMAL_FONT

ws3.cell(info_row + 1, 1).value = "Xmin ="
ws3.cell(info_row + 1, 1).font = BOLD_FONT
ws3.cell(info_row + 1, 2).value = x_min

ws3.cell(info_row + 2, 1).value = "Xmax ="
ws3.cell(info_row + 2, 1).font = BOLD_FONT
ws3.cell(info_row + 2, 2).value = x_max

ws3.cell(info_row + 3, 1).value = "R = Xmax - Xmin ="
ws3.cell(info_row + 3, 1).font = BOLD_FONT
ws3.cell(info_row + 3, 2).value = R

ws3.cell(info_row + 4, 1).value = "h (точн.) = R / (1 + 3.32·lg(n)) ="
ws3.cell(info_row + 4, 1).font = BOLD_FONT
ws3.cell(info_row + 4, 2).value = round(h_raw, 4)

ws3.cell(info_row + 5, 1).value = "h (округл.) ="
ws3.cell(info_row + 5, 1).font = BOLD_FONT
ws3.cell(info_row + 5, 2).value = h

# ---------------------------------------------------------------------------
# Лист 4: Интервальный ряд (Таблица 2)
# ---------------------------------------------------------------------------

ws4 = wb.create_sheet("Интервальный ряд")

ws4.merge_cells("A1:G1")
c = ws4["A1"]
c.value = "Таблица 2. Интервальный ряд"
c.font = Font(bold=True, size=14, name="Calibri")
c.alignment = CENTER

headers4 = ["Интервал", "aᵢ (нижн.)", "bᵢ (верхн.)", "Середина xᵢ*", "kᵢ", "wᵢ = kᵢ/n", "wᵢ/h (плотность)"]
widths4 = [18, 12, 12, 14, 8, 14, 18]
for col, (h_text, w) in enumerate(zip(headers4, widths4), 1):
    style_header(ws4.cell(2, col), h_text)
    set_col_width(ws4, col, w)

cum_k4 = 0
for i, iv in enumerate(intervals):
    row = i + 3
    cum_k4 += iv["k"]
    ws4.cell(row, 1).value = f"[{iv['a']}; {iv['b']})"
    style_data(ws4.cell(row, 1))
    style_data(ws4.cell(row, 2), iv["a"])
    style_data(ws4.cell(row, 3), iv["b"])
    style_data(ws4.cell(row, 4), iv["mid"])
    style_data(ws4.cell(row, 5), iv["k"])
    style_data(ws4.cell(row, 6), round(iv["w"], 6))
    style_data(ws4.cell(row, 7), round(iv["density"], 6))
    if i % 2 == 0:
        for col in range(1, 8):
            ws4.cell(row, col).fill = LIGHT_FILL

sum_row4 = len(intervals) + 3
style_sum(ws4.cell(sum_row4, 1), "Σ")
style_sum(ws4.cell(sum_row4, 2), "")
style_sum(ws4.cell(sum_row4, 3), "")
style_sum(ws4.cell(sum_row4, 4), "")
style_sum(ws4.cell(sum_row4, 5), n)
style_sum(ws4.cell(sum_row4, 6), round(sum(iv["w"] for iv in intervals), 6))
style_sum(ws4.cell(sum_row4, 7), "")

info4_row = sum_row4 + 2
ws4.cell(info4_row, 1).value = "Мода (Mo) ="
ws4.cell(info4_row, 1).font = BOLD_FONT
ws4.cell(info4_row, 2).value = round(Mo, 4)
ws4.cell(info4_row, 2).font = BOLD_FONT

ws4.cell(info4_row + 1, 1).value = "Медиана (Me) ="
ws4.cell(info4_row + 1, 1).font = BOLD_FONT
ws4.cell(info4_row + 1, 2).value = round(Me, 4)
ws4.cell(info4_row + 1, 2).font = BOLD_FONT

# ---------------------------------------------------------------------------
# Лист 5: Полигон и F*(x) (Таблица 3 + графики)
# ---------------------------------------------------------------------------

ws5 = wb.create_sheet("Полигон и F(x)")

# Заголовок
ws5.merge_cells("A1:F1")
c = ws5["A1"]
c.value = "Таблица 3. Полигон частот и Эмпирическая функция распределения F*(x)"
c.font = Font(bold=True, size=14, name="Calibri")
c.alignment = CENTER

# Таблица полигона (по точечному ряду)
ws5.merge_cells("A2:C2")
c = ws5["A2"]
c.value = "Полигон частот (по точечному ряду)"
c.font = BOLD_FONT
c.alignment = CENTER
c.fill = SUBHEADER_FILL

headers5a = ["xᵢ", "kᵢ", "wᵢ"]
for col, h_text in enumerate(headers5a, 1):
    style_header(ws5.cell(3, col), h_text)
    set_col_width(ws5, col, 12)

for i, (v, f, w) in enumerate(zip(unique_vals, freqs, rel_freqs)):
    row = i + 4
    style_data(ws5.cell(row, 1), v)
    style_data(ws5.cell(row, 2), f)
    style_data(ws5.cell(row, 3), round(w, 6))
    if i % 2 == 0:
        for col in range(1, 4):
            ws5.cell(row, col).fill = LIGHT_FILL

# Таблица F*(x)
ws5.merge_cells("E2:G2")
c = ws5["E2"]
c.value = "Эмпирическая функция F*(x)"
c.font = BOLD_FONT
c.alignment = CENTER
c.fill = SUBHEADER_FILL

headers5b = ["xᵢ", "F*(x) = Σwⱼ (j≤i)", "Накопл. k"]
for col, h_text in enumerate(headers5b, 5):
    style_header(ws5.cell(3, col), h_text)
    set_col_width(ws5, col, 20)

cum_f = 0
for i, (v, f, w, cdf) in enumerate(zip(unique_vals, freqs, rel_freqs, cdf_vals)):
    row = i + 4
    style_data(ws5.cell(row, 5), v)
    style_data(ws5.cell(row, 6), round(cdf, 6))
    style_data(ws5.cell(row, 7), int(cdf * n))
    if i % 2 == 0:
        for col in range(5, 8):
            ws5.cell(row, col).fill = LIGHT_FILL

data_rows = len(unique_vals)

# --- Полигон частот (LineChart) ---
chart_poly = LineChart()
chart_poly.title = "Полигон частот"
chart_poly.style = 10
chart_poly.y_axis.title = "wᵢ (относит. частота)"
chart_poly.x_axis.title = "xᵢ"
chart_poly.width = 18
chart_poly.height = 12

data_poly = Reference(ws5, min_col=3, min_row=3, max_row=3 + data_rows)
cats_poly = Reference(ws5, min_col=1, min_row=4, max_row=3 + data_rows)
chart_poly.add_data(data_poly, titles_from_data=True)
chart_poly.set_categories(cats_poly)
chart_poly.series[0].graphicalProperties.line.solidFill = "2E75B6"
chart_poly.series[0].marker.symbol = "circle"
chart_poly.series[0].marker.size = 5
ws5.add_chart(chart_poly, "A" + str(data_rows + 6))

# --- Эмпирическая функция F*(x) (LineChart ступенчатый) ---
chart_cdf = LineChart()
chart_cdf.title = "Эмпирическая функция распределения F*(x)"
chart_cdf.style = 10
chart_cdf.y_axis.title = "F*(x)"
chart_cdf.x_axis.title = "x"
chart_cdf.width = 18
chart_cdf.height = 12

data_cdf = Reference(ws5, min_col=6, min_row=3, max_row=3 + data_rows)
cats_cdf = Reference(ws5, min_col=5, min_row=4, max_row=3 + data_rows)
chart_cdf.add_data(data_cdf, titles_from_data=True)
chart_cdf.set_categories(cats_cdf)
chart_cdf.series[0].graphicalProperties.line.solidFill = "ED7D31"
ws5.add_chart(chart_cdf, "J" + str(data_rows + 6))

# --- Гистограмма (BarChart) по интервальному ряду ---
# Вставим данные для гистограммы под таблицей интервального ряда

hist_start_row = data_rows + 6
ws5.merge_cells(f"A{hist_start_row - 2}:D{hist_start_row - 2}")
c = ws5.cell(hist_start_row - 2, 1)
c.value = "Данные для гистограммы (интервальный ряд)"
c.font = BOLD_FONT
c.alignment = CENTER
c.fill = SUBHEADER_FILL

hist_header_row = hist_start_row - 1
style_header(ws5.cell(hist_header_row, 9), "Интервал (середина)")
style_header(ws5.cell(hist_header_row, 10), "wᵢ/h (плотность)")
set_col_width(ws5, 9, 22)
set_col_width(ws5, 10, 18)

for i, iv in enumerate(intervals):
    row = hist_header_row + 1 + i
    style_data(ws5.cell(row, 9), iv["mid"])
    style_data(ws5.cell(row, 10), round(iv["density"], 6))

chart_hist = BarChart()
chart_hist.type = "col"
chart_hist.title = "Гистограмма (плотность частоты)"
chart_hist.style = 10
chart_hist.y_axis.title = "wᵢ/h"
chart_hist.x_axis.title = "Середина интервала"
chart_hist.width = 18
chart_hist.height = 12

hist_data_rows = len(intervals)
data_hist = Reference(ws5, min_col=10, min_row=hist_header_row, max_row=hist_header_row + hist_data_rows)
cats_hist = Reference(ws5, min_col=9, min_row=hist_header_row + 1, max_row=hist_header_row + hist_data_rows)
chart_hist.add_data(data_hist, titles_from_data=True)
chart_hist.set_categories(cats_hist)
chart_hist.series[0].graphicalProperties.solidFill = "2E75B6"
ws5.add_chart(chart_hist, "S" + str(data_rows + 6))

# ---------------------------------------------------------------------------
# Лист 6: Начальные моменты (Таблица 4)
# ---------------------------------------------------------------------------

ws6 = wb.create_sheet("Начальные моменты")

ws6.merge_cells("A1:G1")
c = ws6["A1"]
c.value = "Таблица 4. Начальные моменты"
c.font = Font(bold=True, size=14, name="Calibri")
c.alignment = CENTER

headers6 = ["xᵢ", "kᵢ", "wᵢ = kᵢ/n", "xᵢ·wᵢ", "xᵢ²·wᵢ", "xᵢ³·wᵢ", "xᵢ⁴·wᵢ"]
widths6 = [10, 8, 14, 16, 16, 20, 22]
for col, (h_text, w) in enumerate(zip(headers6, widths6), 1):
    style_header(ws6.cell(2, col), h_text)
    set_col_width(ws6, col, w)

sum_xw = sum_x2w = sum_x3w = sum_x4w = sum_w = 0
for i, (v, f, w) in enumerate(zip(unique_vals, freqs, rel_freqs)):
    row = i + 3
    xw = v * w
    x2w = v**2 * w
    x3w = v**3 * w
    x4w = v**4 * w
    sum_xw += xw
    sum_x2w += x2w
    sum_x3w += x3w
    sum_x4w += x4w
    sum_w += w
    style_data(ws6.cell(row, 1), v)
    style_data(ws6.cell(row, 2), f)
    style_data(ws6.cell(row, 3), round(w, 6))
    style_data(ws6.cell(row, 4), round(xw, 6))
    style_data(ws6.cell(row, 5), round(x2w, 6))
    style_data(ws6.cell(row, 6), round(x3w, 4))
    style_data(ws6.cell(row, 7), round(x4w, 2))
    if i % 2 == 0:
        for col in range(1, 8):
            ws6.cell(row, col).fill = LIGHT_FILL

sum_row6 = len(unique_vals) + 3
style_sum(ws6.cell(sum_row6, 1), "Σ")
style_sum(ws6.cell(sum_row6, 2), n)
style_sum(ws6.cell(sum_row6, 3), round(sum_w, 6))
style_sum(ws6.cell(sum_row6, 4), round(sum_xw, 6))
style_sum(ws6.cell(sum_row6, 5), round(sum_x2w, 6))
style_sum(ws6.cell(sum_row6, 6), round(sum_x3w, 4))
style_sum(ws6.cell(sum_row6, 7), round(sum_x4w, 2))

info6 = sum_row6 + 2
ws6.cell(info6, 1).value = "α₁ = Σ(xᵢ·wᵢ) ="
ws6.cell(info6, 1).font = BOLD_FONT
ws6.cell(info6, 2).value = round(alpha1, 6)
ws6.cell(info6, 2).font = NORMAL_FONT

ws6.cell(info6 + 1, 1).value = "α₂ = Σ(xᵢ²·wᵢ) ="
ws6.cell(info6 + 1, 1).font = BOLD_FONT
ws6.cell(info6 + 1, 2).value = round(alpha2, 6)

ws6.cell(info6 + 2, 1).value = "α₃ = Σ(xᵢ³·wᵢ) ="
ws6.cell(info6 + 2, 1).font = BOLD_FONT
ws6.cell(info6 + 2, 2).value = round(alpha3, 4)

ws6.cell(info6 + 3, 1).value = "α₄ = Σ(xᵢ⁴·wᵢ) ="
ws6.cell(info6 + 3, 1).font = BOLD_FONT
ws6.cell(info6 + 3, 2).value = round(alpha4, 2)

# ---------------------------------------------------------------------------
# Лист 7: Характеристики выборки
# ---------------------------------------------------------------------------

ws7 = wb.create_sheet("Характеристики выборки")

ws7.merge_cells("A1:D1")
c = ws7["A1"]
c.value = "Характеристики выборки"
c.font = Font(bold=True, size=14, name="Calibri")
c.alignment = CENTER

rows7 = [
    ("Характеристика", "Формула", "Значение", ""),
    ("Выборочное среднее x̄в = α₁",
     "x̄в = Σ(xᵢ·wᵢ)",
     round(x_mean, 6), ""),
    ("Выборочная дисперсия Dв",
     "Dв = α₂ − α₁²",
     round(D_v, 6), ""),
    ("Среднеквадр. откл. σв",
     "σв = √Dв",
     round(sigma_v, 6), ""),
    ("", "", "", ""),
    ("Центральный момент β₂ = Dв",
     "β₂ = α₂ − α₁²",
     round(beta2, 6), ""),
    ("Центральный момент β₃",
     "β₃ = α₃ − 3α₁α₂ + 2α₁³",
     round(beta3, 6), ""),
    ("Центральный момент β₄",
     "β₄ = α₄ − 4α₁α₃ + 6α₁²α₂ − 3α₁⁴",
     round(beta4, 6), ""),
    ("", "", "", ""),
    ("Асимметрия A",
     "A = β₃ / σв³",
     round(A, 6), ""),
    ("Эксцесс E",
     "E = β₄ / σв⁴ − 3",
     round(E, 6), ""),
]

set_col_width(ws7, 1, 42)
set_col_width(ws7, 2, 34)
set_col_width(ws7, 3, 16)
set_col_width(ws7, 4, 10)

for i, row_data in enumerate(rows7):
    row = i + 2
    for col, val in enumerate(row_data, 1):
        cell = ws7.cell(row, col)
        cell.value = val
        cell.border = cell_border()
        if i == 0:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
        elif val == "":
            cell.fill = PatternFill("solid", fgColor="F2F2F2")
        else:
            cell.font = BOLD_FONT if col == 1 else NORMAL_FONT
            cell.alignment = CENTER if col == 3 else LEFT

# ---------------------------------------------------------------------------
# Лист 8: Точечные оценки (п.7)
# ---------------------------------------------------------------------------

ws8 = wb.create_sheet("Точечные оценки")

ws8.merge_cells("A1:D1")
c = ws8["A1"]
c.value = "Точечные оценки параметров (п.7)"
c.font = Font(bold=True, size=14, name="Calibri")
c.alignment = CENTER

set_col_width(ws8, 1, 40)
set_col_width(ws8, 2, 34)
set_col_width(ws8, 3, 16)
set_col_width(ws8, 4, 10)

rows8 = [
    ("Параметр", "Формула", "Значение", ""),
    ("--- 7.1 Метод статистических оценок ---", "", "", ""),
    ("x̄г = x̄в",
     "x̄г = x̄в = α₁",
     round(x_mean, 6), ""),
    ("Dг = S² (исправленная дисперсия)",
     "S² = n/(n−1) · Dв",
     round(S2, 6), ""),
    ("σг = S (исправл. СКО)",
     "S = √(n/(n−1)) · σв",
     round(S, 6), ""),
    ("--- 7.2 Метод моментов ---", "", "", ""),
    ("x̄г = x̄в",
     "α₁ = α₁' ⇒ x̄г = x̄в",
     round(x_mean, 6), ""),
    ("Dг = Dв",
     "β₂ = β₂ ⇒ Dг = Dв",
     round(D_v, 6), ""),
    ("σг = σв",
     "σг = σв",
     round(sigma_v, 6), ""),
]

for i, row_data in enumerate(rows8):
    row = i + 2
    for col, val in enumerate(row_data, 1):
        cell = ws8.cell(row, col)
        cell.value = val
        cell.border = cell_border()
        if i == 0:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
        elif isinstance(val, str) and val.startswith("---"):
            cell.font = BOLD_FONT
            cell.fill = SUBHEADER_FILL
            cell.alignment = LEFT
            if col == 1:
                ws8.merge_cells(f"A{row}:D{row}")
                break
        else:
            cell.font = BOLD_FONT if col == 1 else NORMAL_FONT
            cell.alignment = CENTER if col == 3 else LEFT

# ---------------------------------------------------------------------------
# Лист 9: Интервальные оценки (п.8)
# ---------------------------------------------------------------------------

ws9 = wb.create_sheet("Интервальные оценки")

ws9.merge_cells("A1:D1")
c = ws9["A1"]
c.value = "Интервальные оценки (п.8), γ = 0.95"
c.font = Font(bold=True, size=14, name="Calibri")
c.alignment = CENTER

set_col_width(ws9, 1, 44)
set_col_width(ws9, 2, 34)
set_col_width(ws9, 3, 20)
set_col_width(ws9, 4, 10)

rows9 = [
    ("Параметр", "Формула / Пояснение", "Значение", ""),
    ("--- 8.1 Дов. интервал для мат. ожидания ---", "", "", ""),
    ("Уровень доверия γ", "", 0.95, ""),
    ("t (квантиль при γ=0.95)", "Стандартное: t=1.96", 1.96, ""),
    ("S (исправл. СКО)", "S = √(n/(n−1)·Dв)", round(S, 6), ""),
    ("Δ = t·S/√n", f"Δ = {t}·{round(S,4)}/√{n}", round(delta, 6), ""),
    (f"Нижн. граница: x̄г − Δ",
     f"{round(x_mean,4)} − {round(delta,4)}",
     round(mean_low, 6), ""),
    (f"Верхн. граница: x̄г + Δ",
     f"{round(x_mean,4)} + {round(delta,4)}",
     round(mean_high, 6), ""),
    ("Интервал для x̄г",
     f"({round(mean_low,4)}; {round(mean_high,4)})",
     "", ""),
    ("--- 8.2 Дов. интервал для σ ---", "", "", ""),
    ("q (для n=120, γ=0.95)", "q = 0.143", 0.143, ""),
    ("Нижн. граница: S·(1−q)",
     f"{round(S,4)}·(1−{q})",
     round(sigma_low, 6), ""),
    ("Верхн. граница: S·(1+q)",
     f"{round(S,4)}·(1+{q})",
     round(sigma_high, 6), ""),
    ("Интервал для σг",
     f"({round(sigma_low,4)}; {round(sigma_high,4)})",
     "", ""),
]

for i, row_data in enumerate(rows9):
    row = i + 2
    for col, val in enumerate(row_data, 1):
        cell = ws9.cell(row, col)
        cell.value = val
        cell.border = cell_border()
        if i == 0:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
        elif isinstance(val, str) and val.startswith("---"):
            cell.font = BOLD_FONT
            cell.fill = SUBHEADER_FILL
            cell.alignment = LEFT
            if col == 1:
                ws9.merge_cells(f"A{row}:D{row}")
                break
        else:
            cell.font = BOLD_FONT if col == 1 else NORMAL_FONT
            cell.alignment = CENTER if col == 3 else LEFT

# ---------------------------------------------------------------------------
# Сохранение файла
# ---------------------------------------------------------------------------

output_file = "mathstat_rgr.xlsx"
wb.save(output_file)
print(f"✅ Файл '{output_file}' успешно создан!")
print()
print("=== Краткие результаты ===")
print(f"n = {n},  Xmin = {x_min},  Xmax = {x_max},  R = {R}")
print(f"h = {h}  (точное: {h_raw:.4f})")
print(f"Мода Mo  = {Mo:.4f}")
print(f"Медиана Me = {Me:.4f}")
print(f"x̄в = α₁ = {alpha1:.6f}")
print(f"Dв  = {D_v:.6f}")
print(f"σв  = {sigma_v:.6f}")
print(f"β₃  = {beta3:.6f}")
print(f"β₄  = {beta4:.6f}")
print(f"A (асимметрия) = {A:.6f}")
print(f"E (эксцесс)    = {E:.6f}")
print(f"S²  = {S2:.6f}")
print(f"S   = {S:.6f}")
print(f"Δ   = {delta:.6f}")
print(f"ДИ для x̄: ({mean_low:.4f}; {mean_high:.4f})")
print(f"ДИ для σ:  ({sigma_low:.4f}; {sigma_high:.4f})")
